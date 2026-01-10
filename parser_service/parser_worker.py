#!/usr/bin/env python3
import os
import sys
import asyncio
import redis.asyncio as redis
from pathlib import Path
from datetime import datetime
import traceback
import openpyxl
from async_impl import parse_moex_stock_async, get_investing_price_async

REDIS_HOST = os.getenv('REDIS_HOST', 'redis')
REDIS_PORT = int(os.getenv('REDIS_PORT', 6379))
JOBS_STREAM = 'parser:jobs'
RESULTS_STREAM = 'parser:results'
CONSUMER_GROUP = 'parser_service'
BATCH_SIZE = int(os.getenv('BATCH_SIZE', 10))

redis_client = None


async def get_redis():
    global redis_client
    if redis_client is None:
        redis_client = redis.Redis(
            host=REDIS_HOST,
            port=REDIS_PORT,
            decode_responses=True
        )
    return redis_client


def format_date_for_api(date: datetime) -> str:
    return date.strftime('%Y-%m-%d')


async def process_single_stock_async(row_num: int, stock_name: str, ticker: str, investing_url: str, 
                                    target_date: str, index: int):
    moex_price = None
    num_trades = None
    volume = None
    investing_price = None
    
    if ticker:
        try:
            results = await parse_moex_stock_async(ticker, target_date)
            if results:
                for entry in results:
                    if entry.get('date') == target_date:
                        moex_price = entry.get('close_price')
                        num_trades = entry.get('num_trades')
                        volume = entry.get('volume')
                        break
        except Exception as e:
            print(f"  [{index}] {stock_name} - MOEX error: {e}", file=sys.stderr)
    
    if investing_url and moex_price is not None:
        try:
            investing_price = await get_investing_price_async(investing_url, target_date)
        except Exception as e:
            print(f"  [{index}] {stock_name} - Investing.com error: {e}", file=sys.stderr)
    
    return row_num, stock_name, ticker, moex_price, num_trades, volume, investing_price


async def process_excel_file(file_content: bytes, date: datetime) -> tuple[bytes, str]:
    temp_input = Path(f"/tmp/input_{os.getpid()}.xlsx")
    temp_output = Path(f"/tmp/output_{os.getpid()}.xlsx")
    
    try:
        temp_input.write_bytes(file_content)
        
        print(f"Loading Excel file...")
        wb = openpyxl.load_workbook(temp_input)
        ws = wb.active
        
        ws.cell(1, 4).value = date
        
        if not ws.cell(2, 17).value:
            ws.cell(2, 17).value = "Количество сделок"
        if not ws.cell(2, 18).value:
            ws.cell(2, 18).value = "Объем торгов"
        
        target_date = format_date_for_api(date)
        
        stocks_data = []
        row_num = 4
        
        while True:
            isin = ws.cell(row_num, 2).value
            if not isin:
                break
            
            stock_name = ws.cell(row_num, 3).value
            investing_url = ws.cell(row_num, 14).value
            ticker = ws.cell(row_num, 15).value
            
            stocks_data.append({
                'row_num': row_num,
                'stock_name': stock_name,
                'ticker': ticker,
                'investing_url': investing_url
            })
            
            row_num += 1
        
        total_rows = len(stocks_data)
        successful_moex = 0
        successful_investing = 0
        
        print(f"\nProcessing {total_rows} stocks for date: {date.strftime('%d.%m.%Y')}")
        print(f"Using batch size: {BATCH_SIZE} concurrent requests")
        print("-" * 80)
        
        for batch_start in range(0, total_rows, BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, total_rows)
            batch = stocks_data[batch_start:batch_end]
            
            print(f"\nProcessing batch {batch_start//BATCH_SIZE + 1}/{(total_rows + BATCH_SIZE - 1)//BATCH_SIZE} "
                  f"(stocks {batch_start + 1}-{batch_end})...")
            
            tasks = [
                process_single_stock_async(
                    stock['row_num'],
                    stock['stock_name'],
                    stock['ticker'],
                    stock['investing_url'],
                    target_date,
                    batch_start + i + 1
                )
                for i, stock in enumerate(batch)
            ]
            
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    print(f"  [{batch_start + i + 1}] Error: {result}", file=sys.stderr)
                    continue
                
                row_num, stock_name, ticker, moex_price, num_trades, volume, investing_price = result
                
                print(f"  [{batch_start + i + 1}] {stock_name} ({ticker})")
                
                if moex_price is not None:
                    ws.cell(row_num, 5).value = moex_price
                    ws.cell(row_num, 19).value = num_trades if num_trades is not None else 0
                    ws.cell(row_num, 20).value = volume if volume is not None else 0
                    print(f"    MOEX: ✓ {moex_price} RUB (trades: {num_trades}, vol: {volume})")
                    successful_moex += 1
                else:
                    print(f"    MOEX: ✗ Not found")
                
                if investing_price is not None:
                    ws.cell(row_num, 6).value = investing_price
                    print(f"    Investing.com: ✓ ${investing_price}")
                    successful_investing += 1
                else:
                    print(f"    Investing.com: ✗ Not found")
        
        print("\n" + "=" * 80)
        summary = (
            f"📊 Summary:\n"
            f"  Total stocks: {total_rows}\n"
            f"  MOEX prices found: {successful_moex}/{total_rows}\n"
            f"  Investing.com prices found: {successful_investing}/{total_rows}"
        )
        print(summary)
        
        print(f"\nSaving results...")
        wb.save(temp_output)
        
        result_content = temp_output.read_bytes()
        
        return result_content, summary
    
    finally:
        temp_input.unlink(missing_ok=True)
        temp_output.unlink(missing_ok=True)


async def process_job(job_data: dict):
    job_id = job_data['job_id']
    user_id = job_data['user_id']
    filename = job_data['filename']
    date_str = job_data['date']
    file_content = bytes.fromhex(job_data['file_content'])
    
    print(f"\n{'='*80}")
    print(f"📋 Processing job: {job_id}")
    print(f"👤 User: {user_id}")
    print(f"📁 File: {filename}")
    print(f"📅 Date: {date_str}")
    print(f"{'='*80}\n")
    
    try:
        date = datetime.strptime(date_str, '%d.%m.%Y')
        
        result_content, summary = await process_excel_file(file_content, date)
        
        r = await get_redis()
        result_data = {
            'job_id': job_id,
            'user_id': user_id,
            'status': 'success',
            'filename': filename,
            'file_content': result_content.hex(),
            'summary': summary
        }
        
        await r.xadd(RESULTS_STREAM, result_data)
        print(f"\n✅ Job {job_id} completed successfully!")
    
    except Exception as e:
        print(f"\n❌ Job {job_id} failed with error: {e}")
        traceback.print_exc()
        
        r = await get_redis()
        error_data = {
            'job_id': job_id,
            'user_id': user_id,
            'status': 'error',
            'error': str(e)
        }
        
        await r.xadd(RESULTS_STREAM, error_data)


async def main():
    r = await get_redis()
    
    try:
        await r.xgroup_create(JOBS_STREAM, CONSUMER_GROUP, id='0', mkstream=True)
        print(f"✅ Created consumer group: {CONSUMER_GROUP}")
    except redis.ResponseError as e:
        if 'BUSYGROUP' in str(e):
            print(f"ℹ️  Consumer group already exists: {CONSUMER_GROUP}")
        else:
            raise
    
    print(f"🚀 Parser service started!")
    print(f"👂 Listening for jobs on stream: {JOBS_STREAM}")
    print(f"🔧 Batch size: {BATCH_SIZE}")
    print(f"{'='*80}\n")
    
    while True:
        try:
            messages = await r.xreadgroup(
                CONSUMER_GROUP,
                f'worker-{os.getpid()}',
                {JOBS_STREAM: '>'},
                count=1,
                block=1000
            )
            
            for stream, stream_messages in messages:
                for message_id, job_data in stream_messages:
                    try:
                        await process_job(job_data)
                        await r.xack(JOBS_STREAM, CONSUMER_GROUP, message_id)
                    except Exception as e:
                        print(f"Error processing job: {e}")
                        traceback.print_exc()
        
        except Exception as e:
            print(f"Error reading from Redis stream: {e}")
            await asyncio.sleep(5)


if __name__ == '__main__':
    asyncio.run(main())

