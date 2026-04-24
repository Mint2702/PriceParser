#!/usr/bin/env python3
import os
import sys
import asyncio
import logging
import redis.asyncio as redis
from pathlib import Path
from datetime import datetime
import traceback
import openpyxl
import xml.etree.ElementTree as ET
from curl_cffi.requests import AsyncSession

from async_impl import get_investing_price_async


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

REDIS_HOST = os.getenv('REDIS_HOST', 'redis')
REDIS_PORT = int(os.getenv('REDIS_PORT', 6379))
JOBS_STREAM = 'us_parser:jobs'
RESULTS_STREAM = 'us_parser:results'
CONSUMER_GROUP = 'us_parser_service'
BATCH_SIZE = int(os.getenv('BATCH_SIZE', 5))
INVESTING_MAX_RETRIES = int(os.getenv('INVESTING_MAX_RETRIES', 3))
INVESTING_RETRY_DELAY = float(os.getenv('INVESTING_RETRY_DELAY', 2.0))

redis_client = None


async def get_redis():
    global redis_client
    if redis_client is None:
        redis_client = redis.Redis(
            host=REDIS_HOST,
            port=REDIS_PORT,
            decode_responses=True
        )
    return redis_client


def format_date_for_api(date: datetime) -> str:
    return date.strftime('%Y-%m-%d')


def normalize_price(price) -> float | None:
    if price is None:
        return None

    if isinstance(price, (int, float)):
        return float(price)

    if isinstance(price, str):
        price = price.strip()

        if ',' in price:
            price = price.replace('.', '')
            price = price.replace(',', '.')
        elif price.count('.') > 1:
            price = price.replace('.', '')

        try:
            return float(price)
        except ValueError:
            return None

    return None


async def get_usd_rate_from_cbr(date: datetime) -> float | None:
    try:
        date_str = date.strftime('%d/%m/%Y')
        url = f"https://cbr.ru/scripts/XML_daily.asp?date_req={date_str}"

        async with AsyncSession() as client:
            response = await client.get(url, timeout=30, impersonate="chrome120")
            response.raise_for_status()

            root = ET.fromstring(response.content)

            for valute in root.findall('Valute'):
                char_code = valute.find('CharCode')
                if char_code is not None and char_code.text == 'USD':
                    value = valute.find('Value')
                    if value is not None and value.text:
                        usd_rate = float(value.text.replace(',', '.'))
                        return usd_rate

            return None
    except Exception as e:
        logger.error(f"Error fetching USD rate from CBR: {e}")
        return None


async def process_single_stock_async(row_num: int, stock_name: str, investing_url: str,
                                     target_date: str, index: int):
    investing_price = None

    if investing_url:
        for attempt in range(INVESTING_MAX_RETRIES):
            try:
                investing_price = await get_investing_price_async(investing_url, target_date)
                if investing_price is not None:
                    break
                if attempt < INVESTING_MAX_RETRIES - 1:
                    delay = INVESTING_RETRY_DELAY * (2 ** attempt)
                    logger.warning(
                        f"  [{index}] {stock_name} - Investing.com returned None, "
                        f"retrying in {delay}s (attempt {attempt + 1}/{INVESTING_MAX_RETRIES})"
                    )
                    await asyncio.sleep(delay)
            except Exception as e:
                if attempt < INVESTING_MAX_RETRIES - 1:
                    delay = INVESTING_RETRY_DELAY * (2 ** attempt)
                    logger.warning(
                        f"  [{index}] {stock_name} - Investing.com error: {e}, "
                        f"retrying in {delay}s (attempt {attempt + 1}/{INVESTING_MAX_RETRIES})"
                    )
                    await asyncio.sleep(delay)
                else:
                    logger.error(
                        f"  [{index}] {stock_name} - Investing.com error after "
                        f"{INVESTING_MAX_RETRIES} attempts: {e}"
                    )

    return row_num, stock_name, investing_price


async def process_excel_file(file_content: bytes, date: datetime, reparse_mode: bool = False) -> tuple[bytes, str]:
    temp_input = Path(f"/tmp/us_input_{os.getpid()}.xlsx")
    temp_output = Path(f"/tmp/us_output_{os.getpid()}.xlsx")

    try:
        temp_input.write_bytes(file_content)

        logger.info(f"Loading Excel file...")
        wb = openpyxl.load_workbook(temp_input)
        ws = wb.active

        ws.cell(1, 4).value = date.strftime('%d.%m.%Y')

        target_date = format_date_for_api(date)

        logger.info(f"Fetching USD exchange rate from CBR...")
        usd_rate = await get_usd_rate_from_cbr(date)
        if usd_rate:
            logger.info(f"USD rate: {usd_rate} RUB")
        else:
            logger.warning(f"Could not fetch USD rate from CBR")

        stocks_data = []
        row_num = 4

        while True:
            isin = ws.cell(row_num, 2).value
            if not isin:
                break

            col_e = ws.cell(row_num, 5).value

            if reparse_mode:
                if col_e != "ERROR":
                    row_num += 1
                    continue

            stock_name = ws.cell(row_num, 3).value
            investing_url = ws.cell(row_num, 14).value

            stocks_data.append({
                'row_num': row_num,
                'stock_name': stock_name,
                'investing_url': investing_url
            })

            row_num += 1

        total_rows = len(stocks_data)
        successful = 0
        error_count = 0

        mode_str = "REPARSE (ERROR rows only)" if reparse_mode else "FULL"
        logger.info(f"\nProcessing {total_rows} stocks for date: {date.strftime('%d.%m.%Y')} [Mode: {mode_str}]")
        logger.info(f"Using batch size: {BATCH_SIZE} concurrent requests")
        logger.info("-" * 80)

        for batch_start in range(0, total_rows, BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, total_rows)
            batch = stocks_data[batch_start:batch_end]

            logger.info(
                f"\nProcessing batch {batch_start // BATCH_SIZE + 1}/"
                f"{(total_rows + BATCH_SIZE - 1) // BATCH_SIZE} "
                f"(stocks {batch_start + 1}-{batch_end})..."
            )

            tasks = [
                process_single_stock_async(
                    stock['row_num'],
                    stock['stock_name'],
                    stock['investing_url'],
                    target_date,
                    batch_start + i + 1
                )
                for i, stock in enumerate(batch)
            ]

            results = await asyncio.gather(*tasks, return_exceptions=True)

            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    logger.error(f"  [{batch_start + i + 1}] Error: {result}")
                    continue

                row_num, stock_name, investing_price = result

                logger.info(f"  [{batch_start + i + 1}] {stock_name}")

                ws.cell(row_num, 4).value = date.strftime('%d.%m.%Y')

                if investing_price is not None:
                    normalized_price = normalize_price(investing_price)
                    ws.cell(row_num, 5).value = normalized_price
                    logger.info(f"    Investing.com: ✓ ${normalized_price}")
                    successful += 1

                    if usd_rate is not None:
                        ws.cell(row_num, 6).value = usd_rate
                        ws.cell(row_num, 7).value = round(normalized_price * usd_rate, 2)
                elif stocks_data[batch_start + i].get('investing_url'):
                    ws.cell(row_num, 5).value = "ERROR"
                    logger.info(f"    Investing.com: ✗ Not found (ERROR)")
                    error_count += 1
                else:
                    logger.info(f"    Investing.com: ✗ No URL provided")

            await asyncio.sleep(0.3)

        logger.info("\n" + "=" * 80)
        summary = (
            f"📊 Summary:\n"
            f"  Total stocks: {total_rows}\n"
            f"  Investing.com prices found: {successful}/{total_rows}\n"
            f"  ERRORs: {error_count}"
        )
        logger.info(summary)

        logger.info(f"\nSaving results...")
        wb.save(temp_output)

        result_content = temp_output.read_bytes()

        return result_content, summary

    finally:
        temp_input.unlink(missing_ok=True)
        temp_output.unlink(missing_ok=True)


async def process_job(job_data: dict):
    job_id = job_data['job_id']
    user_id = job_data['user_id']
    filename = job_data['filename']
    file_content = bytes.fromhex(job_data['file_content'])
    mode = job_data.get('mode', 'parse')
    reparse_mode = (mode == 'reparse')

    if reparse_mode:
        temp_file = Path(f"/tmp/us_check_{os.getpid()}.xlsx")
        try:
            temp_file.write_bytes(file_content)
            wb = openpyxl.load_workbook(temp_file)
            ws = wb.active
            date_value = ws.cell(1, 4).value

            if isinstance(date_value, datetime):
                date = date_value
                date_str = date.strftime('%d.%m.%Y')
            elif isinstance(date_value, str):
                date_str = date_value
                date = datetime.strptime(date_str, '%d.%m.%Y')
            else:
                raise ValueError("Date not found in Excel file (cell D1)")
        finally:
            temp_file.unlink(missing_ok=True)
    else:
        date_str = job_data['date']
        date = datetime.strptime(date_str, '%d.%m.%Y')

    logger.info(f"\n{'=' * 80}")
    logger.info(f"📋 Processing job: {job_id}")
    logger.info(f"👤 User: {user_id}")
    logger.info(f"📁 File: {filename}")
    logger.info(f"📅 Date: {date_str}")
    logger.info(f"🔄 Mode: {'REPARSE (ERROR rows only)' if reparse_mode else 'FULL'}")
    logger.info(f"{'=' * 80}\n")

    try:
        result_content, summary = await process_excel_file(file_content, date, reparse_mode)

        r = await get_redis()
        result_data = {
            'job_id': job_id,
            'user_id': user_id,
            'status': 'success',
            'filename': filename,
            'file_content': result_content.hex(),
            'summary': summary
        }

        await r.xadd(RESULTS_STREAM, result_data)
        logger.info(f"\n✅ Job {job_id} completed successfully!")

    except Exception as e:
        logger.error(f"\n❌ Job {job_id} failed with error: {e}")
        traceback.print_exc()

        r = await get_redis()
        error_data = {
            'job_id': job_id,
            'user_id': user_id,
            'status': 'error',
            'error': str(e)
        }

        await r.xadd(RESULTS_STREAM, error_data)


async def main():
    r = await get_redis()

    try:
        await r.xgroup_create(JOBS_STREAM, CONSUMER_GROUP, id='0', mkstream=True)
        logger.info(f"✅ Created consumer group: {CONSUMER_GROUP}")
    except redis.ResponseError as e:
        if 'BUSYGROUP' in str(e):
            logger.info(f"ℹ️  Consumer group already exists: {CONSUMER_GROUP}")
        else:
            raise

    logger.info(f"🚀 US Parser service started!")
    logger.info(f"👂 Listening for jobs on stream: {JOBS_STREAM}")
    logger.info(f"🔧 Batch size: {BATCH_SIZE}")
    logger.info(f"{'=' * 80}\n")

    while True:
        try:
            messages = await r.xreadgroup(
                CONSUMER_GROUP,
                f'worker-{os.getpid()}',
                {JOBS_STREAM: '>'},
                count=1,
                block=1000
            )

            for stream, stream_messages in messages:
                for message_id, job_data in stream_messages:
                    try:
                        await process_job(job_data)
                        await r.xack(JOBS_STREAM, CONSUMER_GROUP, message_id)
                    except Exception as e:
                        logger.error(f"Error processing job: {e}")
                        traceback.print_exc()

        except Exception as e:
            logger.error(f"Error reading from Redis stream: {e}")
            await asyncio.sleep(5)


if __name__ == '__main__':
    asyncio.run(main())
