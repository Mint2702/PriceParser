#!/usr/bin/env python3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "parser_service"))

import argparse
import asyncio
import openpyxl
from datetime import datetime
from sync import parse_moex_stock, get_stock_id, get_stock_data
from async_impl import parse_moex_stock_async, get_investing_price_async


def parse_date(date_str: str) -> datetime:
    try:
        return datetime.strptime(date_str, '%d.%m.%Y')
    except ValueError:
        raise ValueError(f"Invalid date format: {date_str}. Expected format: DD.MM.YYYY (e.g., 31.10.2025)")


def format_date_for_api(date: datetime) -> str:
    return date.strftime('%Y-%m-%d')


def normalize_price(price) -> float | None:
    if price is None:
        return None
    
    if isinstance(price, (int, float)):
        return float(price)
    
    if isinstance(price, str):
        price = price.strip()
        
        if ',' in price:
            price = price.replace('.', '')
            price = price.replace(',', '.')
        elif price.count('.') > 1:
            price = price.replace('.', '')
        
        try:
            return float(price)
        except ValueError:
            return None
    
    return None


def run_sync_version(input_path: Path, output_path: Path, date: datetime):
    print(f"Loading Excel file: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    ws.cell(1, 4).value = date
    
    if not ws.cell(2, 17).value:
        ws.cell(2, 17).value = "Количество сделок"
    if not ws.cell(2, 18).value:
        ws.cell(2, 18).value = "Объем торгов"
    
    target_date = format_date_for_api(date)
    
    row_num = 4
    total_rows = 0
    successful_moex = 0
    successful_investing = 0
    
    print(f"\nProcessing stocks for date: {date.strftime('%d.%m.%Y')} [SYNC MODE]")
    print("-" * 80)
    
    while True:
        isin = ws.cell(row_num, 2).value
        if not isin:
            break
        
        total_rows += 1
        stock_name = ws.cell(row_num, 3).value
        investing_url = ws.cell(row_num, 14).value
        ticker = ws.cell(row_num, 15).value
        
        print(f"\n[{total_rows}] {stock_name} ({ticker})")
        
        if ticker:
            print(f"  Fetching MOEX price...", end=' ')
            try:
                results = parse_moex_stock(ticker, target_date)
                moex_price = None
                num_trades = None
                volume = None
                if results:
                    for entry in results:
                        if entry.get('date') == target_date:
                            moex_price = entry.get('close_price')
                            num_trades = entry.get('num_trades')
                            volume = entry.get('volume')
                            break
                
                if moex_price is not None:
                    normalized_price = normalize_price(moex_price)
                    ws.cell(row_num, 5).value = normalized_price
                    ws.cell(row_num, 17).value = num_trades if num_trades is not None else 0
                    ws.cell(row_num, 18).value = volume if volume is not None else 0
                    print(f"✓ {normalized_price} RUB (trades: {num_trades}, vol: {volume})")
                    successful_moex += 1
                else:
                    print("✗ Not found")
            except Exception as e:
                print(f"✗ Error: {e}")
        
        if investing_url:
            print(f"  Fetching Investing.com price...", end=' ')
            try:
                stock_id = get_stock_id(investing_url)
                results = get_stock_data(stock_id, target_date, target_date)
                investing_price = None
                if results:
                    investing_price = results[0].get('close_price')
                
                if investing_price is not None:
                    normalized_price = normalize_price(investing_price)
                    ws.cell(row_num, 6).value = normalized_price
                    print(f"✓ ${normalized_price}")
                    successful_investing += 1
                else:
                    print("✗ Not found")
            except Exception as e:
                print(f"✗ Error: {e}")
        
        row_num += 1
    
    print("\n" + "=" * 80)
    print(f"Summary:")
    print(f"  Total stocks processed: {total_rows}")
    print(f"  MOEX prices found: {successful_moex}/{total_rows}")
    print(f"  Investing.com prices found: {successful_investing}/{total_rows}")
    
    print(f"\nSaving results to: {output_path}")
    wb.save(output_path)
    print("✓ Done!")


async def process_single_stock_async(row_num: int, stock_name: str, ticker: str, investing_url: str, 
                                    target_date: str, index: int):
    moex_price = None
    num_trades = None
    volume = None
    investing_price = None
    
    if ticker:
        try:
            results = await parse_moex_stock_async(ticker, target_date)
            if results:
                for entry in results:
                    if entry.get('date') == target_date:
                        moex_price = entry.get('close_price')
                        num_trades = entry.get('num_trades')
                        volume = entry.get('volume')
                        break
        except Exception as e:
            print(f"  [{index}] {stock_name} - MOEX error: {e}", file=sys.stderr)
    
    if investing_url and moex_price is not None:
        try:
            investing_price = await get_investing_price_async(investing_url, target_date)
        except Exception as e:
            print(f"  [{index}] {stock_name} - Investing.com error: {e}", file=sys.stderr)
    
    return row_num, stock_name, ticker, moex_price, num_trades, volume, investing_price


async def run_async_version(input_path: Path, output_path: Path, date: datetime, batch_size: int = 10):
    print(f"Loading Excel file: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    ws.cell(1, 4).value = date
    
    if not ws.cell(2, 17).value:
        ws.cell(2, 17).value = "Количество сделок"
    if not ws.cell(2, 18).value:
        ws.cell(2, 18).value = "Объем торгов"
    
    target_date = format_date_for_api(date)
    
    stocks_data = []
    row_num = 4
    
    while True:
        isin = ws.cell(row_num, 2).value
        if not isin:
            break
        
        stock_name = ws.cell(row_num, 3).value
        investing_url = ws.cell(row_num, 14).value
        ticker = ws.cell(row_num, 15).value
        
        stocks_data.append({
            'row_num': row_num,
            'stock_name': stock_name,
            'ticker': ticker,
            'investing_url': investing_url
        })
        
        row_num += 1
    
    total_rows = len(stocks_data)
    successful_moex = 0
    successful_investing = 0
    
    print(f"\nProcessing {total_rows} stocks for date: {date.strftime('%d.%m.%Y')} [ASYNC MODE]")
    print(f"Using batch size: {batch_size} concurrent requests")
    print("-" * 80)
    
    for batch_start in range(0, total_rows, batch_size):
        batch_end = min(batch_start + batch_size, total_rows)
        batch = stocks_data[batch_start:batch_end]
        
        print(f"\nProcessing batch {batch_start//batch_size + 1}/{(total_rows + batch_size - 1)//batch_size} "
              f"(stocks {batch_start + 1}-{batch_end})...")
        
        tasks = [
            process_single_stock_async(
                stock['row_num'],
                stock['stock_name'],
                stock['ticker'],
                stock['investing_url'],
                target_date,
                batch_start + i + 1
            )
            for i, stock in enumerate(batch)
        ]
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        for i, result in enumerate(results):
            if isinstance(result, Exception):
                print(f"  [{batch_start + i + 1}] Error: {result}", file=sys.stderr)
                continue
            
            row_num, stock_name, ticker, moex_price, num_trades, volume, investing_price = result
            
            print(f"  [{batch_start + i + 1}] {stock_name} ({ticker})")
            
            if moex_price is not None:
                normalized_price = normalize_price(moex_price)
                ws.cell(row_num, 5).value = normalized_price
                ws.cell(row_num, 17).value = num_trades if num_trades is not None else 0
                ws.cell(row_num, 18).value = volume if volume is not None else 0
                print(f"    MOEX: ✓ {normalized_price} RUB (trades: {num_trades}, vol: {volume})")
                successful_moex += 1
            else:
                print(f"    MOEX: ✗ Not found")
            
            if investing_price is not None:
                normalized_price = normalize_price(investing_price)
                ws.cell(row_num, 6).value = normalized_price
                print(f"    Investing.com: ✓ ${normalized_price}")
                successful_investing += 1
            else:
                print(f"    Investing.com: ✗ Not found")
    
    print("\n" + "=" * 80)
    print(f"Summary:")
    print(f"  Total stocks processed: {total_rows}")
    print(f"  MOEX prices found: {successful_moex}/{total_rows}")
    print(f"  Investing.com prices found: {successful_investing}/{total_rows}")
    
    print(f"\nSaving results to: {output_path}")
    wb.save(output_path)
    print("✓ Done!")


def main():
    parser = argparse.ArgumentParser(
        description='Parse stock prices from MOEX and Investing.com and fill Excel quotations template',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s quotations_template.xlsx 31.10.2025
  %(prog)s input.xlsx 15.12.2025 -o output.xlsx
  %(prog)s input.xlsx 15.12.2025 --async -b 20
        """
    )
    
    parser.add_argument(
        'excel_file',
        type=str,
        help='Path to the Excel file (quotations template)'
    )
    
    parser.add_argument(
        'date',
        type=str,
        help='Date for stock prices in DD.MM.YYYY format (e.g., 31.10.2025)'
    )
    
    parser.add_argument(
        '-o', '--output',
        type=str,
        help='Output file path (default: adds _filled suffix to input filename)',
        default=None
    )
    
    parser.add_argument(
        '--async',
        dest='use_async',
        action='store_true',
        help='Use async mode for faster processing (default: sync mode)'
    )
    
    parser.add_argument(
        '-b', '--batch-size',
        type=int,
        help='Number of concurrent requests per batch (async mode only, default: 10)',
        default=10
    )
    
    args = parser.parse_args()
    
    input_path = Path(args.excel_file)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)
    
    try:
        date = parse_date(args.date)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = input_path.parent / f"{input_path.stem}_filled{input_path.suffix}"
    
    try:
        if args.use_async:
            asyncio.run(run_async_version(input_path, output_path, date, args.batch_size))
        else:
            run_sync_version(input_path, output_path, date)
    except Exception as e:
        print(f"\nError: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
